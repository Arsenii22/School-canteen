import logging
from re import I
from aiogram.types import ReplyKeyboardRemove, ReplyKeyboardMarkup, KeyboardButton, InlineKeyboardMarkup, InlineKeyboardButton
from aiogram import Bot, Dispatcher, executor, types
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup
from fuzzywuzzy import process
from openpyxl import Workbook, load_workbook
import matplotlib.pyplot as plt
import asyncio
import datetime
import sqlite3
import config

database = sqlite3.connect("db.db")
cur = database.cursor()

bot = Bot(token=config.TOKEN)
storage = MemoryStorage()
dp = Dispatcher(bot, storage=storage)

class Form(StatesGroup):
    school = State() # Школа
    like = State() # Нравится еда или нет
    opinion_good = State() # Что больше понравилось / не понравилось
    opinion_bad = State()
    rate = State() # Оценка по 10 балльной шкале

async def get_school_name_by_id(school_id: int):
    cur.execute(f"SELECT name FROM schools WHERE ID={school_id}")
    return cur.fetchall()[0][0]
    
@dp.message_handler(commands="start")
async def start(msg: types.Message):
    if msg.from_user.id in config.MODERATORS_IDS:
        await msg.answer("Приветствуем вас.", reply_markup=ReplyKeyboardRemove())
        await msg.answer("Что вы хотите просмотреть?", reply_markup=InlineKeyboardMarkup().add(InlineKeyboardButton("Топ по школам", callback_data="moderator_top")).add(InlineKeyboardButton("Самые популярные проблемы", callback_data="moderator_top_problems")).add(InlineKeyboardButton("Таблицу с данными по школам", callback_data="moderator_schools_reviews")).add(InlineKeyboardButton("Таблицу со всеми отзывами", callback_data="moderator_all_reviews")))
    else:
        await msg.answer_sticker(r"CAACAgIAAxkBAAEZdOdjXSEK1qpOmYrvcQoD3riSA_4zRgACXgYAAlOx9wNkYNhH9eEsgyoE")
        await msg.answer("Приветствуем вас в нашем чат-боте. В нём вы можете написать рецензию на вашу столовую в школе", reply_markup=ReplyKeyboardMarkup(resize_keyboard=True).add(KeyboardButton("Пройти опрос")))

@dp.message_handler(text="Пройти опрос")
async def review(msg: types.Message):
    cur.execute(f"SELECT timestamp FROM comments WHERE user_id={msg.from_user.id} ORDER BY timestamp DESC LIMIT 1;")
    data = cur.fetchall()
    if data != [] and datetime.datetime.strptime(data[0][0], "%Y-%m-%d %H:%M:%S") + datetime.timedelta(minutes=config.INTERVAL_BETWEEN_REVIEWS) >= datetime.datetime.utcnow():
        await msg.answer("Писать отзывы можно только с периодичностью в 2 минуты")
    
    else:
        await Form.school.set()
        await msg.answer("Не волнуйся, это анонимно :)", reply_markup=ReplyKeyboardRemove())
        await msg.answer("Введите свою школу:")

@dp.callback_query_handler(lambda c: c.data.startswith("moderator_"))
async def moderator(query: types.CallbackQuery):
    await bot.answer_callback_query(query.id)

    if query.data == "moderator_top":
        fig, ax = plt.subplots()
        vals, labels = [], []

        cur.execute(f"SELECT school_id, ROUND(AVG(rate), 2) FROM comments GROUP BY school_id ORDER BY AVG(rate) DESC")
        top_schools = cur.fetchall()
        text = "Топ по школам Калининград:\n"

        for i in range(5):
            text += f"{i + 1}. {await get_school_name_by_id(top_schools[i][0])} ({top_schools[i][1]})\n"
        
        for i in top_schools:
            labels.append(await get_school_name_by_id(i[0]))
            vals.append(i[1])
        
        ax.pie(vals, labels=labels)
        ax.axis("equal")

        plt.savefig("plt.png")
        
        m = await query.message.reply(text)
        await m.reply_photo(types.InputFile("plt.png"))

    elif query.data == "moderator_top_problems":
        fig, ax = plt.subplots()
        vals, labels = [], []

        cur.execute(f"SELECT opinion_bad, COUNT(opinion_bad) FROM comments GROUP BY opinion_bad ORDER BY COUNT(opinion_bad) DESC LIMIT 5;")
        top_problems = cur.fetchall()
        text = "Самые популярные проблемы:\n"

        for i in range(3):
            text += f"{i + 1}. {config.OPINIONS[top_problems[i][0]]} ({top_problems[i][1]})\n"
        
        for i in top_problems:
            labels.append(config.OPINIONS[i[0]])
            vals.append(i[1])

        ax.pie(vals, labels=labels)
        ax.axis("equal")

        plt.savefig("plt.png")
        
        m = await query.message.reply(text)
        await m.reply_photo(types.InputFile("plt.png"))
    
    elif query.data == "moderator_schools_reviews":
        wb = Workbook()
        table = wb.active

        table.append(["Школа", "Сотношение нравиться/не нравиться", "Самые популярные плюсы", "Самые популярные минусы", "Средняя оценка"])

        cur.execute(f"SELECT * FROM schools")

        for i in cur.fetchall():
            cur.execute(f"SELECT ROUND(AVG(like), 2), ROUND(AVG(rate), 2) FROM comments WHERE school_id={i[0]};")
            data = cur.fetchall()[0]

            if None in data:
                table.append([i[1], f"Нет данных", f"Нет данных", f"Нет данных", f"Нет данных"])
            else:
                cur.execute(f"SELECT opinion_bad FROM comments WHERE school_id={i[0]} GROUP BY opinion_bad HAVING COUNT(opinion_bad)=(SELECT COUNT(opinion_bad) FROM comments WHERE school_id={i[0]} GROUP BY opinion_bad ORDER BY COUNT(opinion_bad) DESC LIMIT 1)")
                data_opinion_bad = [config.OPINIONS[i[0]] for i in cur.fetchall()]
                cur.execute(f"SELECT opinion_good FROM comments WHERE school_id={i[0]} GROUP BY opinion_good HAVING COUNT(opinion_good)=(SELECT COUNT(opinion_good) FROM comments WHERE school_id={i[0]} GROUP BY opinion_good ORDER BY COUNT(opinion_good) DESC LIMIT 1)")
                data_opinion_good = [config.OPINIONS[i[0]] for i in cur.fetchall()]
                
                table.append([i[1], f"{data[0] * 100}%", ", ".join(data_opinion_bad), ", ".join(data_opinion_good), f"{data[1]}"])

        wb.save(f"Таблица.xlsx")

        await query.message.reply_document(document=types.InputFile("Таблица.xlsx"))
    
    elif query.data == "moderator_all_reviews":
        wb = Workbook()
        table = wb.active

        table.append(["Школа", "Нравиться/не нравиться", "Плюсы", "Минусы", "Оценка", "Метка времени"])

        cur.execute(f"SELECT school_id, like, opinion_good, opinion_bad, rate, timestamp FROM comments")

        for i in cur.fetchall():
            table.append([await get_school_name_by_id(i[0]), bool(i[1]), config.OPINIONS[i[2]], config.OPINIONS[i[3]], i[4], i[5]])

        wb.save(f"Таблица.xlsx")

        await query.message.reply_document(document=types.InputFile("Таблица.xlsx"))
    
    await query.message.answer("Что вы хотите просмотреть?", reply_markup=InlineKeyboardMarkup().add(InlineKeyboardButton("Топ по школам", callback_data="moderator_top")).add(InlineKeyboardButton("Самые популярные проблемы", callback_data="moderator_top_problems")).add(InlineKeyboardButton("Таблицу с данными по школам", callback_data="moderator_schools_reviews")).add(InlineKeyboardButton("Таблицу со всеми отзывами", callback_data="moderator_all_reviews")))

@dp.message_handler(state=Form.school)
async def get_school(msg: types.Message, state: FSMContext):
    cur.execute(f"SELECT name FROM schools")
    best_school = process.extractOne(msg.text, sum(cur.fetchall(), ()))

    if best_school[1] != 0:
        await msg.reply(f"Ваша школа: {best_school[0]}?", reply_markup=InlineKeyboardMarkup().add(InlineKeyboardButton("Да", callback_data="yes")).add(InlineKeyboardButton("Нет", callback_data="no")))
    else:
        await msg.reply(f"Мы не смогли найти вашу школу, попробуйте ещё раз :(")
    
    await asyncio.sleep(10)

@dp.callback_query_handler(state=Form.all_states)
async def answers(query: types.CallbackQuery, state: FSMContext):
    await bot.answer_callback_query(query.id)

    if query.data == "yes" and query.message.text.startswith("Ваша школа:"):
        async with state.proxy() as data:
            data["school"] = query.message.text.replace("Ваша школа: ", "").replace("?", "")
        await Form.next()

        await query.message.edit_reply_markup()
        await query.message.reply("Нравится ли вам еда в школьной столовой?", reply_markup=InlineKeyboardMarkup().add(InlineKeyboardButton("Да", callback_data="yes")).add(InlineKeyboardButton("Нет", callback_data="no")))

    elif query.data == "no" and query.message.text.startswith("Ваша школа:"):
        async with state.proxy() as data:
            data["school"] = "."
        
        await query.message.edit_reply_markup()
        await query.message.edit_text("Введите свою школу:")

    elif query.data == "yes" and query.message.text == "Нравится ли вам еда в школьной столовой?": # КОСТЫЛЬ ПИЗДЕЦ
        async with state.proxy() as data:
            data["like"] = True
        await Form.next()

        await query.message.edit_reply_markup()
        await query.message.reply("Что вам больше всего понравилось?", reply_markup=InlineKeyboardMarkup().add(InlineKeyboardButton("Вкусная еда", callback_data="opinion_tasty")).add(InlineKeyboardButton("Аппетитная еда", callback_data="opinion_appetitno")).add(InlineKeyboardButton("Свежая еда", callback_data="opinion_fresh")).add(InlineKeyboardButton("Сытная еда", callback_data="opinion_satisfying")).add(InlineKeyboardButton("Подача", callback_data="opinion_giving")).add(InlineKeyboardButton("Ничего", callback_data="opinion_nothing")))

    elif (query.data == "no" and query.message.text == "Нравится ли вам еда в школьной столовой?") or (query.data.startswith("opinion_") and query.message.text == "Что вам больше всего понравилось?"): # КОСТЫЛЬ ПИЗДЕЦ
        async with state.proxy() as data:
            data["opinion_good"] = ""
            if query.data.startswith("opinion_"):
                data["opinion_good"] = query.data
            else:
                data["like"] = False
                await Form.next()
        await Form.next()
        
        await query.message.edit_reply_markup()
        await query.message.reply("Что вам больше всего не понравилось?", reply_markup=InlineKeyboardMarkup().add(InlineKeyboardButton("Пересоленая еда", callback_data="opinion_much_salt")).add(InlineKeyboardButton("Пережаренная еда", callback_data="opinion_much_roasted")).add(InlineKeyboardButton("Переваренная еда", callback_data="opinion_much_boiled")).add(InlineKeyboardButton("Сырая еда", callback_data="opinion_raw")).add(InlineKeyboardButton("Невкусная еда", callback_data="opinion_not_tasty")).add(InlineKeyboardButton("Ничего", callback_data="opinion_nothing")))
    elif query.data.startswith("opinion_"):
        async with state.proxy() as data:
            data["opinion_bad"] = query.data
        await Form.next()

        await query.message.edit_reply_markup()
        await query.message.reply("Оцените качество еды по 10 балльной шкале", reply_markup=InlineKeyboardMarkup().row(InlineKeyboardButton("1️⃣", callback_data="1"), InlineKeyboardButton("2️⃣", callback_data="2")).row(InlineKeyboardButton("3️⃣", callback_data="3"), InlineKeyboardButton("4️⃣", callback_data="4")).row(InlineKeyboardButton("5️⃣", callback_data="5"), InlineKeyboardButton("6️⃣", callback_data="6")).row(InlineKeyboardButton("7️⃣", callback_data="7"), InlineKeyboardButton("8️⃣", callback_data="8")).row(InlineKeyboardButton("9️⃣", callback_data="9"), InlineKeyboardButton("🔟", callback_data="10")))
    elif query.data in ["1", "2", "3", "4", "5", "6", "7", "8", "9", "10"]: # ... у меня нет слов
        async with state.proxy() as data:
            data["rate"] = int(query.data)

            cur.execute(f"SELECT ID FROM schools WHERE name='{data['school']}'")
            cur.execute(f"INSERT INTO comments(user_id, school_id, like, opinion_good, opinion_bad, rate, timestamp) VALUES({query.from_user.id}, '{cur.fetchall()[0][0]}', {data['like']}, '{data['opinion_good']}', '{data['opinion_bad']}', {data['rate']}, '{datetime.datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}')")
            database.commit()
        
        await query.message.edit_reply_markup()
        await query.message.reply("Спасибо за прохождение опроса, он обязательно поможет улучшить питание в вашей школьной столовой", reply_markup=ReplyKeyboardMarkup(resize_keyboard=True).add(KeyboardButton("Пройти опрос")))

        await state.finish()
    else:
        await state.finish()

        await query.message.reply("Извините, что-то пошло не так, попробуйте перезапустить бота")

if __name__ == '__main__':
    executor.start_polling(dp, skip_updates=True)